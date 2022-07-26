from asyncio.trsock import TransportSocket
from configparser import MAX_INTERPOLATION_DEPTH
from email.utils import encode_rfc2231
from re import S
import tkinter as tk					
from tkinter import E, N, W, S, Button, ttk
import tkinter
from tracemalloc import take_snapshot
from turtle import bgcolor, onclick
from iapws import IAPWS97
from iapws import iapws97
from tkinter import *
from lzma import FILTER_LZMA2
import math
from iapws import IAPWS97
from iapws import iapws97
from iapws import IAPWS95
import math
import numpy as np
import openpyxl 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill
from openpyxl.styles import colors
root = tk.Tk()
root.title("SIEMENS tools")

'''mygreen = "#d2ffd2"
myred = "#dd0202"

style = ttk.Style()

style.theme_create( "yummy", parent="alt", settings={
        "TNotebook": { "configure": {"tabmargins": [2, 5, 2, 0] }},
        "TNotebook.Tab": {
            "configure": {"padding": [5, 1], "background": mygreen },
            "map":       {"background": [("selected", myred)],
                          "expand": [("selected", [1, 1, 1, 0])]} } })

style.theme_use("yummy")'''
s = ttk.Style()
s.theme_use('default')
s.configure('TNotebook.Tab', background="green3")
s.map("TNotebook", background= [("selected", "green3")])



tabControl = ttk.Notebook(root)

#tab1 = tkinter.Frame(tabControl)
tab2 = tkinter.Frame(tabControl)
tab3 = tkinter.Frame(tabControl)
tab4 = tkinter.Frame(tabControl)
tab5 = tkinter.Frame(tabControl)
#tab6 = tkinter.Frame(tabControl)
tab7 = tkinter.Frame(tabControl)
tab8 = tkinter.Frame(tabControl)
#tab9 = tkinter.Frame(tabControl)
#tab10 = tkinter.Frame(tabControl)
tab11 = tkinter.Frame(tabControl)


#tab1.config(bg="#313132")
#tab1.config(bg="black")
#tab1.config(bg="#404149")
tab3.config(bg="#404149")
tab2.config(bg="#404149")
tab4.config(bg="#404149")
tab5.config(bg="#404149")
#tab6.config(bg="#404149")
tab7.config(bg="#404149")
tab8.config(bg="#404149")
#tab9.config(bg="#404149")
#tab10.config(bg="#404149")
tab11.config(bg="#404149")



label_text = StringVar()

#NPSH
###########################################################################################################
label_text = StringVar()
''''Label(tab1,text="").grid(row=0,sticky=W)
Label(tab1,text="Site elevation",font=("Arial", 11),fg ="green").grid(row=1,column=0)
Label(tab1,text="m",font=("Arial", 11),fg ="green").grid(row=1,column=2)
Label(tab1,text="Friction Losses in Horizontal pipe, fittings and Piping components like NRV, strainer etc. ",font=("bold", 11),fg ="green").grid(row=2,sticky=W)
Label(tab1,text="MWC ",font=("Arial", 11),fg ="green").grid(row=2,column=2)
Label(tab1,text="TG deck Elevation",font=("Arial", 11),fg ="green").grid(row=3,column=0)
Label(tab1,text="M",font=("Arial", 11),fg ="green").grid(row=3,column=2)
Label(tab1,text="Turbine Centerline",font=("Arial", 11),fg ="green").grid(row=4,column=0)
Label(tab1,text="M",font=("Arial", 11),fg ="green").grid(row=4,column=2)
Label(tab1,text="Lube Oil Skid floor Elevation",font=("Arial", 11),fg ="green").grid(row=5,column=0)
Label(tab1,text="M",font=("Arial", 11),fg ="green").grid(row=5,column=2)
Label(tab1,text="Net positive suction head Required as recommended by Shaft driven pump manufacturer. ",font=("Arial", 11),fg ="green").grid(row=6,column=0)
Label(tab1,text="MWC ",font=("Arial", 11),fg ="green").grid(row=6,column=2)
Label(tab1,text="Actual  elevation rasied  of Lube oil system after deducting equipment Pedestal ",font=("Arial", 11),fg ="green").grid(row=7,column=0)
Label(tab1,text="MT ",font=("Arial", 11),fg ="green").grid(row=7,column=2)
Label(tab1,text="").g
#c4= Entry(tab1)
a= Entry(tab1)
#c13 = Entry(tab1)
b = Entry(tab1)
c = Entry(tab1)
d =Entry(tab1)
f = Entry(tab1)
#c22 = Entry(tab1)
h = Entry(tab1)
#c27 = Entry(tab1)
k = Entry(tab1)
a.grid(row=1,column=1)
b.grid(row=2,column=1)
c.grid(row=3,column=1)
d.grid(row=4,column=1)
f.grid(row=5,column=1)
h.grid(row=6,column=1)
k.grid(row=7,column=1)
 
def function1():
    #num = float(a.get())
    l = 101325
    m = 288.16
    n = 9.80665
    o = 0.02896968
    p = 8.314462618
    q = o*n
    r = m*p
    s = -1*(q/r)
    t = s*float(a.get())
    u = math.exp(t)
    v =l*u
    w = v*0.0075
    x = w*13.663/10001
    y = x
    z = -20
    i = -1*z*13.663/1000
    j = y - i
    #num2 = float(c13.get())
    #num3 = 
   #num4 = 
    #num5 = 
    #c17 = float(c.get())+float(d.get())-float(e.get())
    #Label(tab1,text="value of c17 is - "+str(c13.get())).grid(row=14,column=1)
    v = float(c.get())+float(d.get())-float(f.get())
    #Label(tab1,text="jkhj").grid(row=12,column=1)
    #c18 = c17 + float(c13.get())
    a = v + float(b.get())
    Label(tab1,text="value of c17 is - "+str(v)).grid(row=14,column=1)
    var = j-a
    #num6 = float(c22.get())
    vx = float(h.get())+1
    #num7 = float(c27.get())
    vy = float(h.get())
    vz = var +  float(k.get())
    ab = var
    ans = round(ab - vy,2)
    #c38 = c34 - c32
    
    Label(tab1,text="Result are"+str(ans)).grid(row=12,column=1,sticky=W)
b = Button(tab1, text="Calculate", command=function1,bg="blue",fg="white")
b.grid(row=10, column=0, columnspan=2, rowspan=2, sticky=W + E + N + S, padx=5, pady=5)'''
#################################################################


#tabControl.add(tab1, text ='NPSH Calculation ')
tabControl.add(tab3, text ='Turing gear motor sizing  ')
tabControl.add(tab2,text='Pump Power Rating ')
tabControl.add(tab4,text='LS Coupling Interference')
tabControl.add(tab5,text="Volume of Deaerator Btn Normal LLWL(No Dish Ends)")
#tabControl.add(tab6,text="Pressure Drop Calc across Strainer")
tabControl.add(tab7,text="PRDS Calculation")
tabControl.add(tab8,text="Lube oil")
#tabControl.add(tab9,text="NPSH")
#tabControl.add(tab10,text="Pressure Drop Calc across Strainer")
tabControl.add(tab11,text="NPSH Calculation")


tabControl.pack(expand = 1, fill ="both")

# Turing gear motor sizing
#########################################################################################################
'''Label(tab3,text="",fg="#ffff00",bg="#404149").grid(row=0,column=1)
label_text = StringVar()
Label(tab3,text='Turbine',font = "Verdana 10 bold",bg="green").grid(row=0,column=0)

Label(tab3, text="Moment of Inertia (mr2) (J1)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=1, column=0,sticky=W)
Label(tab3, text="kg-m2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=1, column=2,sticky=W)
Label(tab3, text="Breakaway Torque (M1)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2, column=0,sticky=W)
Label(tab3, text="Nm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2, column=2,sticky=W)

Label(tab3,text='Gearbox Pinion shaft + HS Coupling',bg="green",font = "Verdana 10 bold").grid(row=3,column=0)

Label(tab3, text="Bearing Diameter (d) ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4, column=0,sticky=W)
Label(tab3, text="mm ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4, column=2,sticky=W)
Label(tab3, text="Shaft Weight (m) ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=0, sticky=W)
Label(tab3, text="Kg ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=2, sticky=W)
Label(tab3, text="HS Coupling Weight ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=0 ,sticky=W)
Label(tab3, text="Kg ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=2,sticky=W)
Label(tab3, text="Coefficient of Friction (f) ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,sticky=W)
Label(tab3, text="Moment of Inertia of pinion shaft (mr2) (J2) ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=8,column=0,sticky=W)
Label(tab3, text="kg-m2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=8,column=2,sticky=W)
#Label(window, text="Label(window, text="Moment of Inertia of pinion shaft (mr2) (J2) ").grid(row=12,column=0,sticky=W) ").grid(row=12,column=0,sticky=W)
Label(tab3, text="Moment of Inertia of HS Coupling (mr2) (J3)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=9,column=0,sticky=W)
Label(tab3, text="kg-m2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=9,column=2,sticky=W)
Label(tab3, text="Gearbox Pinion shaft Breakaway Torque",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=10,column=0,sticky=W)
Label(tab3, text="Nm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=10,column=2,sticky=W)
Label(tab3, text="",textvariable = label_text ).grid(row=15, column=0, sticky=W)

Label(tab3,text='Gearbox Wheel shaft + Wheel + LS Coupling',bg='green',font = "Verdana 10 bold").grid(row=16,column=0)
Label(tab3,text="",fg="#ffff00",bg="#404149").grid(row=17,column=0)
Label(tab3, text="Bearing Diameter (d)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=18,column=0,sticky=W)
Label(tab3, text="mm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=18,column=2,sticky=W)
Label(tab3, text="Shaft Weight (m)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=19,column=0,sticky=W)
Label(tab3, text="Kg",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=19,column=2,sticky=W)
Label(tab3, text="LS Coupling Weight",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=20,column=0,sticky=W)
Label(tab3, text="Kg",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=20,column=2,sticky=W)
Label(tab3, text="Coefficient of Friction (f)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=21,column=0,sticky=W)
Label(tab3, text="Moment of Inertia of Gear wheel shaft + Wheel (mr2) (J4)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=22,column=0,sticky=W)
Label(tab3, text="kg-m2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=22,column=2,sticky=W)
Label(tab3, text="Moment of Inertia of LS Coupling (mr2) (J5)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=23,column=0,sticky=W)
Label(tab3, text="kg-m2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=23,column=2,sticky=W)
Label(tab3, text="Gearbox Wheel+Wheel Shaft Breakaway Torque",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=24,column=0,sticky=W)
Label(tab3, text="Nm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=24,column=2,sticky=W)

#Label(tab3,text="Generator (with or without jacking oil - project Specific)",bg='yellow').grid(row=0,column=3,sticky=W)
Label(tab3, text="",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=27,column=0,sticky=W)
Label(tab3,text="Moment of Inertia (mr2-J5)",padx=100,font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=1,column=3,sticky=W)
Label(tab3,text="kg-m2",font=("Verdana 10 bold",10),fg="#ffff00",bg="#404149").grid(row=1,column=5,sticky=W)
Label(tab3,text="Breakaway Torque (M4)",padx=100,font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2,column=3,sticky=W)
Label(tab3,text="Nm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2,column=5,sticky=W)
Label(tab3,text = "",fg="#ffff00",bg="#404149").grid(row=3,column=3)

Label(tab3,text="Gearbox Speed Ratio",font = "Verdana 10 bold",bg="yellow").grid(row=4,column=4)
Label(tab3,text = "",fg="#ffff00",bg="#404149").grid(row=5,column=3)
Label(tab3,text = "HS shaft operating speed (n1)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=3)
Label(tab3,text = "rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=5)
Label(tab3,text = "LS shaft operating speed (n2)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=3)
Label(tab3,text = "rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=5)
Label(tab3,text="",fg="#ffff00",bg="#404149").grid(row=8,column=3)

Label(tab3,text="Turning Speed & Motor details",font = "Verdana 10 bold",bg="green").grid(row=9,column=4)
Label(tab3,text="",fg="#ffff00",bg="#404149").grid(row=10,column=3)
Label(tab3,text="Turning motor speed (n3)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=11,column=3)
Label(tab3,text="rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=11,column=5)
Label(tab3,text="HS shaft turning speed (n4)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=12,column=3)
Label(tab3,text="rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=12,column=5)
Label(tab3,text="Ratio of Starting Trq to Nominal Trq of motor at 80% voltage (Tr)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=13,column=3)

i4 = Entry(tab3)
i5 = Entry(tab3)
i9 = Entry(tab3)
i10 = Entry(tab3)
i11 = Entry(tab3)
i12 = Entry(tab3)
i14 = Entry(tab3)
i15 = Entry(tab3)
i17 = Entry(tab3)
i22 = Entry(tab3)
i23 = Entry(tab3)
i24 = Entry(tab3)
i25 = Entry(tab3)
i27 = Entry(tab3)
i28 = Entry(tab3)
i30 = Entry(tab3)
i35 = Entry(tab3)
i36 = Entry(tab3)
i39 = Entry(tab3)
i40 = Entry(tab3)
i44 = Entry(tab3)
i45 = Entry(tab3)
i47 = Entry(tab3)
i4.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i5.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i9.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i10.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i11.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i12.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i14.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i15.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i17.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i22.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i23.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i24.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i25.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i27.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i28.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i30.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i35.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i36.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i39.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i40.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i44.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i45.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i47.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")

i4.grid(row=1, column=1)
i5.grid(row=2, column=1)
i9.grid(row=4, column=1)
i10.grid(row=5, column=1)
i11.grid(row=6, column=1)
i12.grid(row=7, column=1)
i14.grid(row=8, column=1)
i15.grid(row=9,column=1)
i17.grid(row=10,column=1)
i22.grid(row=18,column=1)
i23.grid(row=19,column=1)
i24.grid(row=20,column=1)
i25.grid(row=21,column=1)
i27.grid(row=22,column=1)
i28.grid(row=23,column=1)
i30.grid(row=24,column=1)
i35.grid(row=1,column=4)
i36.grid(row = 2,column=4)
i39.grid(row=6,column=4)
i40.grid(row=7,column=4)
i44.grid(row=11,column=4)
i45.grid(row=12,column=4)
i47.grid(row=13,column=4)

def add_numbers():
    num1 = float(i4.get())
    num2 = float(i5.get())
    num3 = float(i9.get())
    num4 = float(i10.get())
    num5 = float(i11.get())
    num6 = float(i12.get())
    num7 = float(i14.get())
    num8 = float(i15.get())
    num9 = i17.get()
    #num9 = float(self.i17.get())
    num10 = float(i22.get())
    num11 = float(i23.get())
    num12 = float(i24.get())
    num13 = float(i25.get())
    num14 = float(i27.get())
    num15 = float(i28.get())
    num16 = i30.get()
    num17 = float(i35.get())
    num18 = float(i36.get())
    num19 = float(i39.get())
    num20 = float(i40.get())
    num21 = float(i44.get())
    num22 = float(i45.get())
    num23 = float(i47.get())
    ou13 = (num3/float(2000))*((num4 + num5) * float(9.81))*num6
        #a = str(num9) + ""
    ou18 = 0
    if num9 == "" :
        ou18 = ou13
    else:
        a = float(num9)
        ou18 = a

    ou26 = (num10 /2000) *((num11 + num12) * 9.81) * num13
    ou31 = 0
    #b = str(num16) + ""
    if num16 == "" :
        ou31 = ou26
    else :
        b = float(num16)
        ou31 = b

    ou41 =  num19 / num20
    ou46 = num21 / num22    
    ou49 = num2 + ou18 + ((ou31 + num18)/ ou41)
    ou50 = ou49/ou46
    ou51 = (2 * math.pi  * num21 * ou50)/(60000 * num23)
    s = np.array([3.7, 4, 5.5, 11, 15, 18.5, 22, 30, 37, 45])
    ou52 = 0
    j=0
    for i in s :
        if i >= 1.3 * ou51 :
            ou52 = i
            j = j+1
            break
    
    #self.o52.insert(END, str(ou52))
    ou55 = (num1 + num7 +  num8 +( num14 + num15 + num17 )/ou41 /ou41)/ou46/ou46
    ou56 = 2 * math.pi * num21/60
    ou57 = ou52/2/math.pi/num21*60000
    ou58 = ou57/ou55
    ou59 = ou56/ou58
    Label(tab3,text="",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=30,column=0)
    Label(tab3,text="",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=31,column=0)
    if j==0 :
        Label(tab3,text="Motor not availavile",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=32,column=0)
    else :
        Label(tab3,text="Final Motor rating  is :",font="Verdana 14 bold",fg="#ffff00",bg="#404149").grid(row=32,column=0)
        Label(tab3,text=str(ou52)+" "+"Kw",font="Verdana 14 bold",fg="#ffff00",bg="#404149").grid(row=32,column=1)
      
b = Button(tab3, text="Calculate", command=add_numbers,bg="darkblue",fg="white",font=("Verdana 10 bold",13))
b.grid(row=29, column=2, columnspan=2, rowspan=2, sticky=W + E + N + S, padx=5, pady=5)
'''
Label(tab3,text="",fg="#ffff00",bg="#404149").grid(row=0,column=1)
label_text = StringVar()
Label(tab3,text='Turbine',font = "Verdana 10 bold",bg="green").grid(row=0,column=0)

Label(tab3, text="Moment of Inertia (mr2) (J1)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=1, column=0,sticky=W)
Label(tab3, text="kg-m2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=1, column=2,sticky=W)
Label(tab3, text="Breakaway Torque (M1)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2, column=0,sticky=W)
Label(tab3, text="Nm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2, column=2,sticky=W)

Label(tab3,text='Gearbox Pinion shaft + HS Coupling',bg="green",font = "Verdana 10 bold").grid(row=3,column=0)

Label(tab3, text="Bearing Diameter (d) ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4, column=0,sticky=W)
Label(tab3, text="mm ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4, column=2,sticky=W)
Label(tab3, text="Shaft Weight (m) ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=0, sticky=W)
Label(tab3, text="Kg ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=2, sticky=W)
Label(tab3, text="HS Coupling Weight ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=0 ,sticky=W)
Label(tab3, text="Kg ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=2,sticky=W)
Label(tab3, text="Coefficient of Friction (f) ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,sticky=W)
Label(tab3, text="Moment of Inertia of pinion shaft (mr2) (J2) ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=8,column=0,sticky=W)
Label(tab3, text="kg-m2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=8,column=2,sticky=W)
#Label(window, text="Label(window, text="Moment of Inertia of pinion shaft (mr2) (J2) ").grid(row=12,column=0,sticky=W) ").grid(row=12,column=0,sticky=W)
Label(tab3, text="Moment of Inertia of HS Coupling (mr2) (J3)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=9,column=0,sticky=W)
Label(tab3, text="kg-m2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=9,column=2,sticky=W)
Label(tab3, text="Gearbox Pinion shaft Breakaway Torque",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=10,column=0,sticky=W)
Label(tab3, text="Nm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=10,column=2,sticky=W)
Label(tab3, text="",textvariable = label_text ).grid(row=15, column=0, sticky=W)

Label(tab3,text='Gearbox Wheel shaft + Wheel + LS Coupling',bg='green',font = "Verdana 10 bold").grid(row=16,column=0)
Label(tab3,text="",fg="#ffff00",bg="#404149").grid(row=17,column=0)
Label(tab3, text="Bearing Diameter (d)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=18,column=0,sticky=W)
Label(tab3, text="mm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=18,column=2,sticky=W)
Label(tab3, text="Shaft Weight (m)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=19,column=0,sticky=W)
Label(tab3, text="Kg",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=19,column=2,sticky=W)
Label(tab3, text="LS Coupling Weight",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=20,column=0,sticky=W)
Label(tab3, text="Kg",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=20,column=2,sticky=W)
Label(tab3, text="Coefficient of Friction (f)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=21,column=0,sticky=W)
Label(tab3, text="Moment of Inertia of Gear wheel shaft + Wheel (mr2) (J4)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=22,column=0,sticky=W)
Label(tab3, text="kg-m2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=22,column=2,sticky=W)
Label(tab3, text="Moment of Inertia of LS Coupling (mr2) (J5)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=23,column=0,sticky=W)
Label(tab3, text="kg-m2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=23,column=2,sticky=W)
Label(tab3, text="Gearbox Wheel+Wheel Shaft Breakaway Torque",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=24,column=0,sticky=W)
Label(tab3, text="Nm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=24,column=2,sticky=W)

#Label(tab3,text="Generator (with or without jacking oil - project Specific)",bg='yellow').grid(row=0,column=3,sticky=W)
Label(tab3, text="",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=27,column=0,sticky=W)
Label(tab3,text="Moment of Inertia (mr2-J5)",padx=100,font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=1,column=3,sticky=W)
#Label(tab3,text="kg-m2",font=("Verdana 10 bold",10),fg="#ffff00",bg="#404149").grid(row=1,column=5,sticky=W)
Label(tab3,text="Breakaway Torque (M4)",padx=100,font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2,column=3,sticky=W)
#Label(tab3,text="Nm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2,column=5,sticky=W)
Label(tab3,text = "",fg="#ffff00",bg="#404149").grid(row=3,column=3)
#grid(row=4,column=4)
lb = Label(tab3,text="Gearbox Speed Ratio",font = "Verdana 10 bold",bg="green")
lb.place(x=900,y=83)
Label(tab3,text = "",fg="#ffff00",bg="#404149").grid(row=5,column=3)
Label(tab3,text = "HS shaft operating speed (n1)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=3)
#Label(tab3,text = "rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=5)
Label(tab3,text = "LS shaft operating speed (n2)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=3)
#Label(tab3,text = "rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=5)
Label(tab3,text="",fg="#ffff00",bg="#404149").grid(row=8,column=3)
lb1 = Label(tab3,text="Turning Speed & Motor details",font = "Verdana 10 bold",bg="green")
#.grid(row=9,column=4)
lb1.place(x=900,y=205)
Label(tab3,text="",fg="#ffff00",bg="#404149").grid(row=10,column=3)
Label(tab3,text="Turning motor speed (n3)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=11,column=3)
#Label(tab3,text="rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=11,column=5)
Label(tab3,text="HS shaft turning speed (n4)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=12,column=3)
#Label(tab3,text="rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=12,column=5)
l2 = Label(tab3,text="Ratio of Starting Trq to Nominal Trq of motor at 80% voltage (Tr)",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
#.grid(row=13,column=3)
l2.place(x=600,y=283)
i4 = Entry(tab3)
i5 = Entry(tab3)
i9 = Entry(tab3)
i10 = Entry(tab3)
i11 = Entry(tab3)
i12 = Entry(tab3)
i14 = Entry(tab3)
i15 = Entry(tab3)
i17 = Entry(tab3)
i22 = Entry(tab3)
i23 = Entry(tab3)
i24 =Entry(tab3)
i25 = Entry(tab3)
i27 = Entry(tab3)
i28 = Entry(tab3)
i30 = Entry(tab3)
i35 = Entry(tab3)
i36 = Entry(tab3)
i39 = Entry(tab3)
i40 = Entry(tab3)
i44 = Entry(tab3)
i45 = Entry(tab3)
i47 = Entry(tab3)
i4.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i5.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i9.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i10.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i11.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i12.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i14.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i15.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i17.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i22.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i23.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i24.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i25.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i27.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i28.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i30.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i35.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i36.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i39.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i40.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i44.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i45.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
i47.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")

i4.grid(row=1, column=1)
i5.grid(row=2, column=1)
i9.grid(row=4, column=1)
i10.grid(row=5, column=1)
i11.grid(row=6, column=1)
i12.grid(row=7, column=1)
i14.grid(row=8, column=1)
i15.grid(row=9,column=1)
i17.grid(row=10,column=1)
i22.grid(row=18,column=1)
i23.grid(row=19,column=1)
i24.grid(row=20,column=1)
i25.grid(row=21,column=1)
i27.grid(row=22,column=1)
i28.grid(row=23,column=1)
i30.grid(row=24,column=1)
#i35.grid(row=1,column=4)
i35.place(x=1100,y=20)
#i36.grid(row = 2,column=4)
i36.place(x=1100,y=45)
#i39.grid(row=6,column=4)
i39.place(x=1100,y=130)
#i40.grid(row=7,column=4)
i40.place(x=1100,y=160)
#i44.grid(row=11,column=4)
i44.place(x=1100,y=243)
#i45.grid(row=12,column=4)
i45.place(x=1100,y=265)
#i47.grid(row=13,column=4)
i47.place(x=1100,y=287)

a  = Label(tab3,text="kg-m2",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
a.place(x=1290,y=20)
b = Label(tab3,text="Nm",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
b.place(x=1290,y=40)
c = Label(tab3,text="rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
c.place(x=1290,y=125)
d = Label(tab3,text="rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
d.place(x=1290,y=157)
e = Label(tab3,text="rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
e.place(x=1290,y=240)
f = Label(tab3,text="rpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
f.place(x=1290,y=262)

def add_numbers():
    num1 = float(i4.get())
    num2 = float(i5.get())
    num3 = float(i9.get())
    num4 = float(i10.get())
    num5 = float(i11.get())
    num6 = float(i12.get())
    num7 = float(i14.get())
    num8 = float(i15.get())
    num9 = i17.get()
    #num9 = float(self.i17.get())
    num10 = float(i22.get())
    num11 = float(i23.get())
    num12 = float(i24.get())
    num13 = float(i25.get())
    num14 = float(i27.get())
    num15 = float(i28.get())
    num16 = i30.get()
    num17 = float(i35.get())
    num18 = float(i36.get())
    num19 = float(i39.get())
    num20 = float(i40.get())
    num21 = float(i44.get())
    num22 = float(i45.get())
    num23 = float(i47.get())
    ou13 = (num3/float(2000))*((num4 + num5) * float(9.81))*num6
        #a = str(num9) + ""
    ou18 = 0
    if num9 == "" :
        ou18 = ou13
    else:
        a = float(num9)
        ou18 = a

    ou26 = (num10 /2000) *((num11 + num12) * 9.81) * num13
    ou31 = 0
    #b = str(num16) + ""
    if num16 == "" :
        ou31 = ou26
    else :
        b = float(num16)
        ou31 = b

    ou41 =  num19 / num20
    ou46 = num21 / num22    
    ou49 = num2 + ou18 + ((ou31 + num18)/ ou41)
    ou50 = ou49/ou46
    ou51 = (2 * math.pi  * num21 * ou50)/(60000 * num23)
    s = np.array([3.7, 4, 5.5, 11, 15, 18.5, 22, 30, 37, 45])
    ou52 = 0
    j=0
    for i in s :
        if i >= 1.3 * ou51 :
            ou52 = i
            j = j+1
            break
    
    #self.o52.insert(END, str(ou52))
    ou55 = (num1 + num7 +  num8 +( num14 + num15 + num17 )/ou41 /ou41)/ou46/ou46
    ou56 = 2 * math.pi * num21/60
    ou57 = ou52/2/math.pi/num21*60000
    ou58 = ou57/ou55
    ou59 = ou56/ou58
    Label(tab3,text="",font="Verdana 10 bold",fg="green",bg="#404149").grid(row=30,column=0)
    Label(tab3,text="",font="Verdana 10 bold",fg="green",bg="#404149").grid(row=31,column=0)
    if j==0 :
        Label(tab3,text="Motor not availavile",font="Verdana 10 bold",fg="green",bg="#404149").grid(row=32,column=0)
    else :
        Label(tab3,text="Final Motor rating  is :",font="Verdana 15 bold",fg="#ffff00",bg="#404149").grid(row=32,column=0)
        Label(tab3,text=str(ou52)+" "+"Kw",font="Verdana 15 bold",fg="#ffff00",bg="#404149").grid(row=32,column=1)
   
b = Button(tab3, text="Calculate", command=add_numbers,bg="darkblue",fg="white",font=("Verdana 10 bold",13))
b.grid(row=29, column=2, columnspan=2, rowspan=2, sticky=W + E + N + S, padx=5, pady=5)

#Pump Power rating
#########################################################################################################


Label(tab2,text="",fg="#ffff00",bg="#404149").grid(row=0,column=0,sticky=W)
label_text = StringVar()
Label(tab2,text="",fg="#ffff00",bg="#404149").grid(row=1,column=0,sticky=W)
Label(tab2,text = " Pump Power Rating ",bg = 'green',font = "Verdana 10 bold").grid(row=1,column=0,sticky=W)

Label(tab2,text="",fg="#ffff00",bg="#404149").grid(row=2,column=0,sticky= W)
Label(tab2,text="Flow",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=0,sticky = W)
Label(tab2,text="m3/hr",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=2,sticky=W)
Label(tab2,text="Head",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4,column=0,sticky = W)
Label(tab2,text="mlc",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4,column=2,sticky=W)
Label(tab2,text="Efficiency",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=0,sticky=W)

Flow = Entry(tab2)
Head = Entry(tab2)
Efficiency = Entry(tab2)
Flow.grid(row =3,column=1,sticky=W)
Head.grid(row=4,column=1,sticky=W)
Efficiency.grid(row=5,column=1,sticky=W)
Flow.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
Head.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
Efficiency.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")

def function3() :
    res = round(float(Flow.get())*float(Head.get())/float(367)/float(Efficiency.get()),3)
    Label(tab2,text="",bg="#404149").grid(row=8,column=0)
    Label(tab2,text="Power is :",font="Verdana 10 bold",bg="green").grid(row=9,column=0)
    Label(tab2,text=str(res)+" "+" "+"KW",font="Verdana 10 bold",bg="green").grid(row=9,column=1)
    
    #B7 = sheet1["B7"]
    #B7.font = Font(size=13, color='green', bold=True, italic=True)
'''def excel() :
    sheet1["A2"] = "Pressure (bar) "
    a = sheet1["A2"]
    a.font = Font(size=20, underline='single', color='FFBB00', bold=True, italic=True)
    #sheet1["C2"] = Pressure.get()
    #sheet1["A4"] = "Saturation Temprature "
    #workbook.save(filename="Book.xlsx")

    sheet1["A3"] = "Flow"
    A3 = sheet1["A3"]
    A3.font = Font(size=13, bold=True, italic=True)
    sheet1["B3"] = Flow.get()
    B3 =  sheet1["B3"]
    B3.font = Font(size=13, bold=True, italic=True)
    sheet1["C3"] = "m3/hr"
    C3 = sheet1["C3"]
    C3.font = Font(size=13, bold=True, italic=True)

    sheet1["A4"] = "Head"
    A4 = sheet1["A4"]
    A4.font = Font(size=13,bold=True, italic=True)
    sheet1["B4"] = Head.get()
    B4 = sheet1["B4"]
    B4.font = Font(size=13, bold=True, italic=True)
    sheet1["C4"] = "mlc"
    C4 = sheet1["C4"]
    C4.font = Font(size=13, bold=True, italic=True)
    sheet1["A5"] = "Efficiency"

    sheet1["A5"] = "Efficiency"
    A5 =  sheet1["A5"]
    A5.font = Font(size=13, bold=True, italic=True)
    sheet1["B5"] = Efficiency.get()
    B5 = sheet1["B5"]
    B5.font = Font(size=13, bold=True, italic=True)

    sheet1["A7"] = "Power is "
    A7 = sheet1["A7"]
    A7.font = Font(size=13, bold=True, italic=True)
    sheet1["C7"] = "KW"
    C7 = sheet1["C7"]
    C7.font = Font(size=13, bold=True, italic=True)'''
    
    

btn = Button(tab2,text = "calculate",command = function3,font="Verdana 10 bold",bg='blue',fg='white')
btn.grid(row=6,column=0)

'''btn1 = Button(tab2,text="Download Excel",command="excel",font="Verdana 10 bold",bg='blue',fg='white')
btn1.grid(row=10,column=0)'''


#LS Coupling Interference
#######################################################################################################
'''workbook4 = load_workbook(filename="tab4.xlsx")
sheet1 = workbook4.active'''
Label(tab4,text="",fg="#ffff00",bg="#404149").grid(row=0,column=0)
Label(tab4,text="",fg="#ffff00",bg="#404149").grid(row=1,column=0)
Label(tab4,text="LS Coupling Interference",bg = 'green',font = "Verdana 10 bold").grid(row=1,column=0)

Label(tab4,text="",fg="#ffff00",bg="#404149").grid(row=2,column=0)

Label(tab4,text="",fg="#ffff00",bg="#404149").grid(row=3,column=0)
Label(tab4,text="Max",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=1)
Label(tab4,text="Min",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=2)

Label(tab4,text="shaft",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4,column=0)
Label(tab4,text="Hole" ,font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=0)

Label(tab4,text="mm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4,column=3)
Label(tab4,text="mm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=3)

Shaftmax = Entry(tab4)
Shaftmin = Entry(tab4)
Holemax = Entry(tab4)
Holemin = Entry(tab4)
Shaftmax.grid(row=4,column=1)
Shaftmin.grid(row=4,column=2)
Holemax.grid(row=5,column=1)
Holemin.grid(row=5,column=2)
Shaftmax.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
Shaftmin.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
Holemax.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
Holemin.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")

def function4() :
    maxinter = float(Shaftmax.get())-float(Holemin.get())
    mininter = float(Shaftmin.get())-float(Holemax.get())
    Label(tab4,text="Max Interference is : ",font="Verdana 10 bold",bg="#404149").grid(row=9,column=0)
    Label(tab4,text= str(maxinter)+" "+" "+"mm",font="Verdana 10 bold",bg="#404149").grid(row=9,column=1)
    Label(tab4,text="Min Interference is : ",font="Verdana 10 bold",bg="#404149").grid(row=10,column=0)
    Label(tab4,text=str(mininter)+" "+" "+"mm",font="Verdana 10 bold",bg="#404149").grid(row=10,column=1)
    '''sheet1["A7"] = "Max Interference is"
    A7 = sheet1["A7"]
    A7.font = Font(size=10, bold=True)
    sheet1["A7"] = "Min Interference is"
    A8 = sheet1["A8"]
    A8.font = Font(size=10, bold=True)
    sheet1["D7"] = str(maxinter)
    D7 = sheet1["D7"]
    D7.font = Font(size=10, bold=True)
    sheet1["D8"] = str(mininter)
    D8 = sheet1["D8"]
    D8.font = Font(size=10, bold=True)
def e4() :
    sheet1["A2"] = "LS Coupling Interference "
    a = sheet1["A2"]
    a.font = Font(size=16, underline='single', color='FFBB00', bold=True)

    sheet1["A4"] = "Shaft"
    A4 = sheet1["A4"]
    A4.font = Font(size=10, bold=True)
    sheet1["A5"] = "Hole"
    A5 = sheet1["A5"]
    A5.font = Font(size=10, bold=True)
    sheet1["D1"] = "Max"
    D1 = sheet1["D1"]
    D1.font = Font(size=10, bold=True)
    sheet1["E1"] = "min"
    E1 = sheet1["E1"]
    E1.font = Font(size=10, bold=True)
    sheet1["D4"] = Shaftmax.get()
    D4 = sheet1["D4"]
    D4.font = Font(size=10, bold=True,color='00FF00')
    sheet1["E4"] = Shaftmin.get()
    E4 = sheet1["E4"]
    E4.font = Font(size=10, bold=True,color='00FF00')
    sheet1["D5"] = Holemax.get()
    D5 = sheet1["D5"]
    D5.font = Font(size=10, bold=True,color='00FF00')
    sheet1["E5"] = Holemin.get()
    E5 = sheet1["E5"]
    E5.font = Font(size=10, bold=True,color='00FF00')
    sheet1["F4"] = "mm"
    F4 = sheet1["F4"]
    F4.font = Font(size=10, bold=True)
    sheet1["F5"] = "mm"
    F5 = sheet1["F5"]
    F5.font = Font(size=10, bold=True)
    workbook4.save(filename="tab4.xlsx")
    Label(tab4,text="Excel file downloaded",bg="#404149").grid(row=13,column=1)'''
   



btn4 = Button(tab4,text="Calculate",command=function4,bg='blue',fg='white',font="Verdana 10 bold")
btn4.grid(row=7,column=1)
Label(tab4,text="",fg="#ffff00",bg="#404149").grid(row=6,column=0)
Label(tab4,text="",fg="#ffff00",bg="#404149").grid(row=8,column=0)

Label(tab4,text="",bg="#404149").grid(row=11,column=1)

'''bte = Button(tab4,text="Download Excel ",command=e4,bg='blue',fg='white',font="Verdana 10 bold")
bte.grid(row=12,column=1)'''


#Volume of Deaerator Btn Normal & LLWL ( No Dish Ends Considered)
#######################################################################################################
'''workbook5 = load_workbook(filename="newtab5.xlsx")
sheet1 = workbook5.active '''
Label(tab5,text="",bg="#404149").grid(row=0,column=0)
Label(tab5,text="",fg="#ffff00",bg="#404149").grid(row=2,column=0)
Label(tab5,text="Volume of Deaerator Btn Normal & LLWL ( No Dish Ends Considered)",bg='green',font="Verdana 10 bold").grid(row=1,column=1)

Label(tab5,text="Deaerator Shell ID                        (m)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=0)
Label(tab5,text="Low Level from Bottom of tank (m)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4,column=0)
Label(tab5,text="Length of storage tank               (m)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=0)
Label(tab5,text="Storage Tank Pressure                (bar)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=0)
Label(tab5,text="Storage tank Temperature        (°C)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=0)
Label(tab5,text="Flow rate                                        (TPH)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=8,column=0)
Label(tab5,text="",fg="#ffff00",bg="#404149").grid(row=9,column=0)

Label(tab5,text="",fg="#ffff00",bg="#404149").grid(row=10,column=0)
Label(tab5,text="Vol of Shell between bottom of shell & Low water level",bg="green",font="Verdana 10 bold").grid(row=11,column=0)
Label(tab5,text="",fg="#ffff00",bg="#404149").grid(row=12,column=0)

Label(tab5,text="",fg="#ffff00",bg="#404149").grid(row=10,column=1)
Label(tab5,text="Vol of Shell between Top of shell & normal Level",bg="green",font="Verdana 10 bold").grid(row=11,column=1)
Label(tab5,text="",fg="#ffff00",bg="#404149").grid(row=12,column=1)

DSI = Entry(tab5)
LLfB = Entry(tab5)
Length = Entry(tab5)
TankP = Entry(tab5)
TankT = Entry(tab5)
flow_rate = Entry(tab5)

DSI.grid(row=3,column=1)
LLfB.grid(row=4,column=1)
Length.grid(row=5,column=1)
TankP.grid(row=6,column=1)
TankT.grid(row=7,column=1)
flow_rate.grid(row=8,column=1)
DSI.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
LLfB.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
Length.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
TankP.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
TankT.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
flow_rate.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")

def v1():
    x = float(DSI.get())
    Radius = float(DSI.get())/2
    Normallevel = float(DSI.get())*float(0.666667)
    Angeltheta = math.acos((float(Radius)-float(LLfB.get()))/float(Radius))
    Volume = float(Radius)*float(Radius)*float(Length.get())*(float(Angeltheta)-math.sin(Angeltheta)*math.cos(Angeltheta))
    vround = round(Volume,3)
    Label(tab5,text="",bg="#404149").grid(row=14,column=0)
    Label(tab5,text="Volume V1 is "+" "+str(vround)+" "+"cu.m",font="Verdana 10 bold",bg="green").grid(row=15,column=0)
    '''sheet1["D13"] = str(vround)
    D13 = sheet1["D13"]
    D13.font = Font(size=10, bold=True,color='00FF00')'''
def v2():
    Radius = float(DSI.get())/2
    Distance = float(DSI.get())-float(DSI.get())*float(0.666667)
    Angeltheta = math.acos((float(Radius)-float(Distance))/float(Radius))
    Volume = float(Radius)*float(Radius)*float(Length.get())*(float(Angeltheta)-math.sin(Angeltheta)*math.cos(Angeltheta))
    vround = round(Volume,3)
    Label(tab5,text="",bg="#404149").grid(row=14,column=1)
    Label(tab5,text="Volume V2 is "+" "+str(vround)+" "+"cu.m",font="Verdana 10 bold",bg="green").grid(row=15,column=1)
    '''sheet1["D17"] = str(vround)
    D17 = sheet1["D17"]
    D17.font = Font(size=10, bold=True,color='00FF00')'''
def volumetank() :
    Radius = float(DSI.get())/2
    res = float(3.14)*float(Radius)*float(Radius)*float(Length.get())
    res1 = round(res,3)
    Label(tab5,text="",bg="#404149").grid(row=14,column=2)
    Label(tab5,text="Total volume of tank is "+" "+str(res1)+" "+" cu.m",font="Verdana 10 bold",bg="green").grid(row=15,column=2)
    '''sheet1["D19"] = str(res1)
    D19 = sheet1["D19"]
    D19.font = Font(size=10, bold=True,color='00FF00')'''
def retentiontime() :
    Radius = float(DSI.get())/2
    totalV = float(3.14)*float(Radius)*float(Radius)*float(Length.get())
    #v1

    Normallevel1 = float(DSI.get())*float(0.666667)
    Angeltheta1 = math.acos((float(Radius)-float(LLfB.get()))/float(Radius))
    v1 = float(Radius)*float(Radius)*float(Length.get())*(float(Angeltheta1)-math.sin(Angeltheta1)*math.cos(Angeltheta1))
    #v2
    Distance = float(DSI.get())-float(DSI.get())*float(0.666667)
    Angeltheta2 = math.acos((float(Radius)-float(Distance))/float(Radius))
    v2 = float(Radius)*float(Radius)*float(Length.get())*(float(Angeltheta2)-math.sin(Angeltheta2)*math.cos(Angeltheta2))
    
    VolumeNL = float(totalV)-float(v2)-float(v1)
    Label(tab5,text="").grid(row=17,column=3)
    
    x = IAPWS95(P = float(TankP.get())*float(0.1), T = float(TankT.get())+float(273))
    y= x.rho
    VO = float(flow_rate.get())*float(1000)/y
    RT = round((VolumeNL*60)/VO,3)
    Label(tab5,text="",bg="#404149").grid(row=19,column=0)
    Label(tab5,text="Retention Time is "+" "+str(RT)+" "+"minute",font="Verdana 10 bold",bg="green").grid(row=20,column=1)
    '''sheet1["D21"] = str(RT)
    D21 = sheet1["D21"]
    D21.font = Font(size=10, bold=True,color='00FF00')
def e() :
    sheet1["A2"] = "Volume of Deaerator Btn Normal & LLWL ( No Dish Ends Considered) "
    a = sheet1["A2"]
    a.font = Font(size=16, underline='single', color='FFBB00', bold=True)
    sheet1["A4"] = "Deaerator Shell ID "
    A4 = sheet1["A4"]
    A4.font = Font(size=10, bold=True)
    sheet1["A5"] = "Low Level from Bottom of tank"
    A5 = sheet1["A5"]
    A5.font = Font(size=10, bold=True)
    sheet1["A6"] = "Length of storage tank"
    A6 = sheet1["A6"]
    A6.font = Font(size=10, bold=True)
    sheet1["A7"] = "Storage Tank Pressure "
    A7 = sheet1["A7"]
    A7.font = Font(size=10, bold=True)
    sheet1["A8"] = "Storage tank Temperature  "
    A8 = sheet1["A8"]
    A8.font = Font(size=10, bold=True)
    sheet1["A9"] = "Flow rate           "
    A9 = sheet1["A9"]
    A9.font = Font(size=10, bold=True)
    sheet1["D4"] = DSI.get()
    D4 = sheet1["D4"]
    D4.font = Font(size=10, bold=True,color='00FF00')
    sheet1["D5"] = LLfB.get()
    D5 = sheet1["D5"]
    D5.font = Font(size=10, bold=True,color='00FF00')
    sheet1["D6"] = Length.get()
    D6 = sheet1["D6"]
    D6.font = Font(size=10, bold=True,color='00FF00')
    sheet1["D7"] = TankP.get()
    D7 = sheet1["D7"]
    D7.font = Font(size=10, bold=True,color='00FF00')
    sheet1["D8"] = TankT.get()
    D8 = sheet1["D8"]
    D8.font = Font(size=10, bold=True,color='00FF00')
    sheet1["D9"] = flow_rate.get()
    D9 = sheet1["D9"]
    D9.font = Font(size=10, bold=True,color='00FF00')
    sheet1["D10"] = " "
    D10 = sheet1["D10"]
    D10.font = Font(size=10, bold=True,color='00FF00')
    sheet1["A11"] = "Vol of Shell between bottom of shell & Low water level"
    A11 = sheet1["A11"]
    A11.font = Font(size=13, color='FFBB00', bold=True)
    sheet1["A13"] = "volume v1 is"
    A13 = sheet1["A13"]
    A13.font = Font(size=10, bold=True)
    sheet1["A15"] = "Vol of Shell between Top of shell & normal Level"
    A15 = sheet1["A15"]
    A15.font = Font(size=13, color='FFBB00', bold=True)
    sheet1["A17"] = "volume v2 is"
    A17 = sheet1["A17"]
    A17.font = Font(size=10, bold=True)
    sheet1["A19"] = "Total Volume of Tank is "
    A19 = sheet1["A19"]
    A19.font = Font(size=10, bold=True)
    sheet1["A21"] = "Retention Time is  "
    A21 = sheet1["A21"]
    A21.font = Font(size=10, bold=True)
    sheet1["E4"] = "m"
    E4 = sheet1["E4"]
    E4.font = Font(size=10, bold=True)
    sheet1["E5"] = "m"
    E5 = sheet1["E5"]
    E5.font = Font(size=10, bold=True)
    sheet1["E6"] = "m"
    E6 = sheet1["E6"]
    E6.font = Font(size=10, bold=True)
    sheet1["E7"] = "bar"
    E7 = sheet1["E7"]
    E7.font = Font(size=10, bold=True)
    sheet1["E8"] = "°C"
    E8 = sheet1["E8"]
    E8.font = Font(size=10, bold=True)
    sheet1["E9"] = "tph"
    E9 = sheet1["E9"]
    E9.font = Font(size=10, bold=True)
    sheet1["E13"] = "cu.m"
    E13 = sheet1["E13"]
    E13.font = Font(size=10, bold=True)
    sheet1["E17"] = "cu.m"
    E17 = sheet1["E17"]
    E17.font = Font(size=10, bold=True)
    sheet1["E19"] = "cu.m"
    E19 = sheet1["E19"]
    E19.font = Font(size=10, bold=True)
    sheet1["E21"] = "Minute"
    E21 = sheet1["E21"]
    E21.font = Font(size=10, bold=True)
    workbook5.save(filename="newtab5.xlsx")
    Label(tab5,text="",bg="#404149").grid(row=22,column=0)
    Label(tab5,text="Excel File Downloaded ! ",font="Verdana 10 bold",bg="green").grid(row=23,column=2)'''


btn51 = Button(tab5,text="Calculate Volume V1 ",command=v1,font="Verdana 10 bold",bg="blue",fg="white")
btn51.grid(row=13,column=0)

btn52 = Button(tab5,text="Calculate Volume V2",command=v2,font="Verdana 10 bold",bg="blue",fg="white")
btn52.grid(row=13,column=1)

btn53 = Button(tab5,text="Total Volume of Tank",command=volumetank,font="Verdana 10 bold",bg="blue",fg="white")
btn53.grid(row=13,column=2)

'''eb = Button(tab5,text="Download Excel",command=e,font="Verdana 10 bold",bg="blue",fg="white")
eb.grid(row=21,column=2)'''


Label(tab5,text="",fg="#ffff00",bg="#404149").grid(row=14,column=0)
Label(tab5,text="",fg="#ffff00",bg="#404149").grid(row=15,column=0)
Label(tab5,text="",fg="#ffff00",bg="#404149").grid(row=16,column=0)
Label(tab5,text="",fg="#ffff00",bg="#404149").grid(row=17,column=0)

btn54 = Button(tab5,text=" Calculate Retention time",command=retentiontime,font="Verdana 10 bold",bg="blue",fg="white")
btn54.grid(row=18,column=1)
#Pressure Drop Calc across Strainer
#######################################################################################################
'''Label(tab6,text="",fg="#ffff00",bg="#404149").grid(row=0,column=0,sticky=W)
#Label(tab6,text="A.",font="Verdana 10 bold",bg="yellow").grid(row=2,column=0,sticky=W)
Label(tab6,text="",fg="#ffff00",bg="#404149").grid(row=1,column=0,sticky=W)

Label(tab6,text="Outer Dia. Of Pipe (D)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2,column=0,sticky=W)
Label(tab6,text="mm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2,column=2,sticky=W) 
Label(tab6,text="Pipe Thickness (t)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=0,sticky=W)
Label(tab6,text="mm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=2,sticky=W) 
Label(tab6,text="",fg="#ffff00",bg="#404149").grid(row=4,column=0,sticky=W)
#Label(tab6,text="B. Filtration area of Strainer Screen (𝑨𝒇)",font="Verdana 10 bold",bg="yellow").grid(row=5,column=0,sticky=W)
Label(tab6,text="Outer Dia. Of Strainer Screen (Ds)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=0,sticky=W)
Label(tab6,text="mm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=2,sticky=W)
Label(tab6,text="Length of Strainer screen (Ls)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=0,sticky=W)
Label(tab6,text="mm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=2,sticky=W)
Label(tab6,text="Mesh Size:",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=0,sticky=W)
Label(tab6,text="Perforations:",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=8,column=0,sticky=W)
Label(tab6,text="mm",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=8,column=2,sticky=W)
Label(tab6,text="Porosity of ø 6 mm Perforations (PP)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=9,column=0,sticky=W)
Label(tab6,text="%",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=9,column=2,sticky=W)
Label(tab6,text="Open Area for Filtration of 40 Mesh (Ao)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=10,column=0,sticky=W)
Label(tab6,text="%",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=10,column=2,sticky=W)
Label(tab6,text="",fg="#ffff00",bg="#404149").grid(row=11,column=0,sticky=W)
#Label(tab6,text="D.",font="Verdana 10 bold",bg="yellow").grid(row=12,column=0,sticky=W)
Label(tab6,text="Flow rate through strainer (Q)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=12,column=0,sticky=W)
Label(tab6,text="tph",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=12,column=2,sticky=W)
Label(tab6,text="Operating Pressure ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=13,column=0,sticky=W)
Label(tab6,text="ata",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=13,column=2,sticky=W)
Label(tab6,text="Operating Temperature",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=14,column=0)
Label(tab6,text="DegC",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=14,column=2,sticky=W)
Label(tab6,text="Flow Coefficient (Cv) (from manufacturer)",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=15,column=0,sticky=W)

c6 = Entry(tab6)
c7 =  Entry(tab6)
c12 =  Entry(tab6)
c13 =  Entry(tab6)
c14 =  Entry(tab6)
c15 =  Entry(tab6)
c16 =  Entry(tab6)
c17 =  Entry(tab6)
c24 =  Entry(tab6)
c25 =  Entry(tab6)
c26 =  Entry(tab6)
c30 = Entry(tab6)

c6.grid(row=2,column=1,sticky=W)
c7.grid(row=3,column=1,sticky=W)
c12.grid(row=5,column=1,sticky=W)
c13.grid(row=6,column=1,sticky=W)
c14.grid(row=7,column=1,sticky=W)
c15.grid(row=8,column=1,sticky=W)
c16.grid(row=9,column=1,sticky=W)
c17.grid(row=10,column=1,sticky=W)
c24.grid(row=12,column=1,sticky=W)
c25.grid(row=13,column=1,sticky=W)
c26.grid(row=14,column=1,sticky=W)
c30.grid(row=15,column=1,sticky=W)
c6.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c7.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c12.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c13.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c14.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c15.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c16.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c17.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c24.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c25.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c26.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c30.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
def function1():
    #Label(tab6,text="Pressure Drop across strainer at clean condition "+str(c6.get()),font="Verdana 10 bold",bg="green").grid(row=19,column=2,sticky=W)
    c8=float(c6.get())-float(2)*float(c7.get())
    c9 =float(3.14)/4*c8*c8
    c18  =float(3.14)*float(c12.get())*float(c13.get())*float(c16.get())*float(c17.get())/10000
    c20=float(c18/c9)
    y = IAPWS95(P = float(c25.get())*float(0.980665)*float(0.1), T = float(c26.get())+float(273))
    c27= y.rho
    c28 = float(c24.get())*float(1000)/float(c27)
    c29 = c28*3.67
    x = (float(c29)/float(c30.get()))
    c31 = float(x*x)
    c32 = float(c31)/float(14.223343)
  
    Label(tab6,text="Pressure Drop across strainer at clean condition "+str(c31),font="Verdana 10 bold",bg="green").grid(row=19,column=2,sticky=W)
    Label(tab6,text=str(c31)+" "+"psi",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=19,column=3,sticky=W)
    Label(tab6,text=str(c32)+" "+"kg/cm2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=20,column=3,sticky=W)

def function2():
    c8=float(c6.get())-2*float(c7.get())
    c9 =3.14/4*c8*c8
    c18  =3.14*float(c12.get())*float(c13.get())*float(c16.get())*float(c17.get())/10000
    c20=c18/c9
    y = IAPWS95(P = float(c25.get())*float(0.980665)*float(0.1), T = float(c26.get())+float(273))
    c27= y.rho
    c28 = float(c24.get())*1000/c27
    c29 = c28*3.67
    x = (c29/float(c30.get()))
    c31 = x*x
    c32 = c31/14.223343
    y=(c29/(float(c30.get())/2))
    c34 = y*y
    c35 = c34/14.223343
    Label(tab6,text="",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=24,column=2,sticky=W)
    Label(tab6,text="Pressure Drop across strainer at 50% clogged condition",font="Verdana 10 bold",bg="green").grid(row=25,column=2,sticky=W)
    Label(tab6,text=str(c34)+" "+"psi",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=25,column=3,sticky=W)
    Label(tab6,text=str(c35)+" "+"kg/cm2",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=26,column=3,sticky=W)

Label(tab6,text="",fg="#ffff00",bg="#404149").grid(row=16,column=0,sticky=W)
btn61 = Button(tab6,text="Pressure Drop across strainer at clean condition  ",command=function1,font="Verdana 10 bold",bg="blue",fg="white")
btn61.grid(row=17,column=2,sticky=W)
Label(tab6,text="",fg="#ffff00",bg="#404149").grid(row=18,column=0,sticky=W)
Label(tab6,text="",fg="#ffff00",bg="#404149").grid(row=19,column=0,sticky=W)
Label(tab6,text="",fg="#ffff00",bg="#404149").grid(row=20,column=0,sticky=W)
Label(tab6,text="",fg="#ffff00",bg="#404149").grid(row=21,column=0,sticky=W)

btn62 = Button(tab6,text="Pressure Drop across strainer at 50% clogged condition",command=function2,font="Verdana 10 bold",bg="blue",fg="white")
btn62.grid(row=23,column=2,sticky=W)'''

#GSS PRDS Calculation
###############################################################################################################
Label(tab7,text="",fg="#ffff00",bg="#404149").grid(row=0,column=0)
Label(tab7,text="GSS PRDS Calculation""",font="Verdana 10 bold",bg="green").grid(row=1,column=1,sticky=W)
Label(tab7,text="",fg="#ffff00",bg="#404149").grid(row=2,column=1,sticky=W)
Label(tab7,text="PCV Inlet Pressure of Steam ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=1,sticky=W)
Label(tab7,text="ATA",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=3,sticky=W)
Label(tab7,text="PCV Inlet Temperature of Steam",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4,column=1,sticky=W)
Label(tab7,text="Deg C",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4,column=3,sticky=W)
Label(tab7,text="PCV Outlet Pressure of Steam ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=1,sticky=W)
Label(tab7,text="ATA",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=3,sticky=W)
Label(tab7,text="Inlet Pressure of Water",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=1,sticky=W)
Label(tab7,text="ATA",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=3,sticky=W)
Label(tab7,text="Inlet Temperature of Water",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=1,sticky=W)
Label(tab7,text="Deg C",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=3,sticky=W)
Label(tab7,text="DSH outlet Pressure",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=8,column=1,sticky=W)
Label(tab7,text="ATA",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=8,column=3,sticky=W)
Label(tab7,text="DSH Outlet Temperature",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=9,column=1,sticky=W)
Label(tab7,text="Deg C",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=9,column=3,sticky=W)
Label(tab7,text="Steam to be supplied to rear gland",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=10,column=1,sticky=W)
Label(tab7,text="Kg/s",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=10,column=3,sticky=W)
Label(tab7,text="Flow Margin",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=9,column=4,padx=100)
Label(tab7,text="Steam to be supplied to front gland",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=11,column=1,sticky=W)
Label(tab7,text="Kg/s",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=11,column=3,sticky=W)



G26 = Entry(tab7)
G27 = Entry(tab7)
G30 = Entry(tab7)
G34 = Entry(tab7)
G35 = Entry(tab7)
G38 = Entry(tab7)
G39 = Entry(tab7)
steam = Entry(tab7)
margin = Entry(tab7)
gland = Entry(tab7)
G26.grid(row=3,column=2,sticky=W)
G27.grid(row=4,column=2,sticky=W)
G30.grid(row=5,column=2,sticky=W)
G34.grid(row=6,column=2,sticky=W)
G35.grid(row=7,column=2,sticky=W)
G38.grid(row=8,column=2,sticky=W)
G39.grid(row=9,column=2,sticky=W)
steam.grid(row=10,column=2,sticky=W)
margin.grid(row=10,column=4)
gland.grid(row=11,column=2,sticky=W)
G26.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
G27.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
G30.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
G34.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
G35.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
G38.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
G39.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
steam.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
margin.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
gland.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")

def function7() :
  p = float(G26.get())*0.980665*0.1
  t = float(G27.get()) + float(273)
  x = IAPWS95(P = p , T = t)
  G28 = x.h
  G31 = G28
  y = IAPWS95(h=G31,P = float(G30.get())*0.980665*0.1)
  G32 = y.T
  Z = IAPWS95(P = float(G34.get())*0.980665*0.1, T = float(G35.get())+float(273))
  G36 = Z.h
  A = IAPWS95(P = float(G38.get())*0.980665*0.1 , T = float(G39.get())+float(273))
  G40 = A.h
  if margin.get() == "" :
   G42 = round(float(steam.get())*3.6, 3)
  else :
   G42 = round(float(steam.get())*3.6*float(margin.get()), 3)

  G43 = round(G42*(G40-G31)/(G36-G31),3)
  G44 = round(G42 - G43,3)
  if margin.get() == "" :
      G45 = round(float(gland.get())*3.6,3)
  else :
      G45 = round(float(gland.get())*3.6*float(margin.get()),3)

  G46 = round(G44 + G45,3)
  Label(tab7,text="",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=13,column=1)
  Label(tab7,text="",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=14,column=1)
  #Label(tab7,text="Steam to be supplied to rear gland is ",font="Verdana 10 bold",fg="green").grid(row=15,column=1,sticky=W)
  #Label(tab7,text=str(G42)+" "+"tph",font="Verdana 10 bold").grid(row=15,column=2,sticky = W,padx =20)
  Label(tab7,text="Amount of Water to be added is ",font="Verdana 10 bold",bg="#404149").grid(row=16,column=1,sticky=W)
  Label(tab7,text=str(G43)+" "+"tph",font="Verdana 10 bold",bg="#404149").grid(row=16,column=2)
  Label(tab7,text="Inlet Steam to Desuperheater",font="Verdana 10 bold",bg="#404149").grid(row=17,column=1,sticky=W)
  Label(tab7,text=str(G44)+" "+"tph",font="Verdana 10 bold",bg="#404149").grid(row=17,column=2)
  #Label(tab7,text="Steam to be supplied to front gland",font="Verdana 10 bold",fg="green").grid(row=18,column=1,sticky=W)
  #Label(tab7,text=str(G45)+" "+"tph",font="Verdana 10 bold").grid(row=18,column=2)
  Label(tab7,text="Total steam to PCV",font="Verdana 10 bold",bg="#404149").grid(row =19,column=1,sticky=W)
  Label(tab7,text=str(G46)+" "+"tph",font="Verdana 10 bold",bg="#404149").grid(row=19,column=2)

Label(tab7,text="",fg="#ffff00",bg="#404149").grid(row=10,column=1,sticky=W)
Label(tab7,text="",fg="#ffff00",bg="#404149").grid(row=11,column=1,sticky=W)
b7 = Button(tab7,text="Calculate",command=function7,font="Verdana 10 bold",bg="blue",fg="white")
b7.grid(row=12,column=1,sticky=W)

#Lube Oil
####################################################################
class f1:
    def __init__(self,win):
        self.n4 = Label(win,text="LUBE OIL",font="Verdana 10 bold",bg="green")
        self.n4.place(x=8,y=10)
        self.i5a =Label(win,text="Turbine",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.a5i = Entry(tab8)
        self.a5i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i5a.place(x=0,y=35)
        self.a5i.place(x=83,y=35)
        self.r5 =Label(win,text="lpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.r5.place(x=250,y=35)
        self.i6a= Label(win,text="Gear Box",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.a6i = Entry(tab8)
        self.a6i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i6a.place(x=0,y=60)
        self.a6i.place(x=83,y=60)
        self.r6 = Label(win, text="lpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.r6.place(x=250, y=60)
        self.i7a = Label(win, text="Alternator",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.a7i = Entry(tab8)
        self.a7i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i7a.place(x=0, y=85)
        self.a7i.place(x=83, y=85)
        self.r7 = Label(win, text="lpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.r7.place(x=250, y=85)
        self.i8a = Label(win, text="Jacking Oil",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.a8i = Entry(tab8)
        self.a8i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i8a.place(x=0, y=110)
        self.a8i.place(x=83, y=110)
        self.r8 = Label(win, text="lpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.r8.place(x=250, y=110)
        self.n11 = Label(win,text="CONTROL OIL(NORMAL)",font="Verdana 10 bold",bg="green")
        self.n11.place(x=8,y=135)
        self.i17a = Label(win, text="Total",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.a17i =Entry(tab8)
        self.a17i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i17a.place(x=0,y=160)
        self.a17i.place(x=83,y=160)
        self.r8 = Label(win, text="lpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.r8.place(x=250, y=160)
        self.n11 = Label(win, text="Control Oil (Transition)",font="Verdana 10 bold",bg="green")
        self.n11.place(x=8, y=185)
        self.i25a = Label(win, text="Total",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.a25i = Entry(tab8)
        self.a25i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i25a.place(x=0, y=210)
        self.a25i.place(x=83, y=210)
        self.r25 = Label(win, text="lpm",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.r25.place(x=250, y=210)
        self.n11 = Label(win, text="POWER LOSS",font="Verdana 10 bold",bg="green")
        self.n11.place(x=300, y=10)
        self.i34b = Label(win, text="Turbine",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.b34i = Entry(tab8)
        self.b34i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i34b.place(x=310, y=35)
        self.b34i.place(x=520, y=35)
        self.rb34 = Label(win, text="kW",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.rb34.place(x=710, y=35)
        self.i35b = Label(win, text="Gear Box",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.b35i = Entry(tab8)
        self.b35i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i35b.place(x=310, y=60)
        self.b35i.place(x=520, y=60)
        self.rb35 = Label(win, text="kW",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.rb35.place(x=710, y=60)
        self.i36b = Label(win, text="Alternator",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.b36i = Entry(tab8)
        self.b36i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i36b.place(x=310, y=85)
        self.b36i.place(x=520, y=85)
        self.rb36 = Label(win, text="kW",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.rb36.place(x=710, y=85)
        self.n52 = Label(win, text="Oil Tank heater Sizing",font="Verdana 10 bold",bg="green")
        self.n52.place(x=300, y=110)
        self.i52j = Label(win, text="Oil Tank Rundown Volume",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.j52i = Entry(tab8)
        self.j52i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i52j.place(x=310, y=135)
        self.j52i.place(x=520, y=135)
        self.rb34 = Label(win, text="liters",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.rb34.place(x=710, y=135)
        self.i55j1 = Label(win, text="Oil Temperature Difference1",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.j55i1 = Entry(tab8)
        self.j55i1.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i55j1.place(x=310, y=160)
        self.j55i1.place(x=520, y=160)
        self.i55j2 = Label(win, text="Oil Temperature Difference2",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.j55i2 = Entry(tab8)
        self.j55i2.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i55j2.place(x=310, y=185)
        self.j55i2.place(x=520, y=185)
        self.i56j = Label(win, text="Oil Tank Weight",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.j56i = Entry(tab8)
        self.j56i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i56j.place(x=310, y=210)
        self.j56i.place(x=520, y=210)
        self.rb34 = Label(win, text="kg",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.rb34.place(x=710, y=210)
        self.i58j = Label(win, text="Heating Time",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.j58i = Entry(tab8)
        self.j58i.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
        self.i58j.place(x=310, y=235)
        self.j58i.place(x=520, y=235)
        self.rb34 = Label(win, text="hrs",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.rb34.place(x=710, y=235)
        self.b1 = Button(win, text="CALCULATE",fg='#ff0', bg='green',height=2,width=15,command=self.add1)
        self.b1.place(x=300, y=350)
    def add1(self):
        '''my_wb = openpyxl.Workbook()
        my_sheet = my_wb.active
        c2 = my_sheet.cell(row=2, column=1)
        c2.value = "Turbine"
        c1 = my_sheet.cell(row=1, column=1)
        c1.value = "LUBE OIL"
        c2i = my_sheet.cell(row=2, column=3)
        a = self.a5i.get()
        c2i.value = a
        c3i = my_sheet['A3']
        c3i.value ="Gear Box"
        c3ii = my_sheet['C3']
        c3ii.value = self.a6i.get()
        c4i = my_sheet['A4']
        c4i.value = "Alternator"
        c4ii = my_sheet['C4']
        c4ii.value = self.a7i.get()


        c5i = my_sheet['A5']
        c5i.value = "Jacking Oil "
        c5ii = my_sheet['C5']
        c5ii.value = self.a8i.get()
        c6i = my_sheet['A6']
        c6i.value = "Control Oil (Normal) --- Total"
        c6ii = my_sheet['C6']
        c6ii.value = self.a17i.get()
        c7i = my_sheet['A7']
        c7i.value = "Control Oil(Transition) --- Total"
        c7ii = my_sheet['C7']
        c7ii.value = self.a25i.get()
        c8i = my_sheet['A8']
        c8i.value = "Turbine"
        c8ii = my_sheet['C8']
        c8ii.value = self.b34i.get()
        c9i = my_sheet['A9']
        c9i.value = "GearBox"
        c9ii = my_sheet['C9']
        c9ii.value = self.b35i.get()


        c10i = my_sheet['A10']
        c10i.value = "Oil Tank Rundown Volume"
        c10ii = my_sheet['C10']
        c10ii.value = self.j52i.get()
        c11i = my_sheet['A11']
        c11i.value = "Oil Temperature Difference(i)"
        c11ii = my_sheet['C11']
        c11ii.value = self.j55i1.get()
        c12i = my_sheet['A12']
        c12i.value = "Oil Temperature Difference (ii)"
        c12ii = my_sheet['C12']
        c12ii.value = self.j55i2.get()
        c13i = my_sheet['A13']
        c13i.value = "Oil Tank Weight"
        c13ii = my_sheet['C13']
        c13ii.value = self.j56i.get()
        c14i = my_sheet['A14']
        c14i.value = "Heating Time"
        c14ii = my_sheet['C14']
        c14ii.value = self.j58i.get()
        c15i = my_sheet['E1']
        c15i.value = "OUTPUT"
        '''






        g9 = float(self.a5i.get()) + float(self.a6i.get()) + float(self.a7i.get()) + float(self.a8i.get())
        g27 = g9 + float(self.a17i.get())
        g28 = g9 + float(self.a25i.get())
        g29 = g28 * 1.1
        self.o29 = Label(tab8,text ="Total Requirement"+ "                    " + str(g29),font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.o29.place(x=100,y=425)
        g30 = 0.4 * (float(self.a5i.get()) + float(self.a6i.get()) + float(self.a7i.get()))
        g37 = float(self.b34i.get()) + float(self.b35i.get()) + float(self.b36i.get())
        g38 = (g37/4.187)*3600
        d3 = g38 * 1.1
        d4 = 48
        d5 = g27
        d6 = d5 * 60 * 0.848722
        d7 = (d3 /(d6 * 0.48338) + d4)
        d8 = 32
        d9 = 8
        d10 = d3/(d9 * 1000)
        d11 = d10 * 1000/60
        self.o_9 = Label(tab8,text = "Cooling Water Requirement" +"   "+  str(d11),font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.o_9.place(x=100, y=450)
        d12 = 5.00
        d13 = d3/(d12 * 1000)
        d14 = d13 * 1000 /60
        d15 = 666.67
        o32 = 11
        o33 = ((0.4 * float(self.a5i.get()) ) + (0.3 * float(self.a6i.get())) + (0.3 * (0.7 * float(self.a6i.get()))) + (0.4 * float(self.a7i.get())))
        o34 = o33 * o32
        self.o_34 = Label(tab8,text = "Tank Capacity required"  + "           " + str(o34),font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.o_34.place(x=100, y= 475)
        o35 = 0
        o36 = o35/o33
        j53 = 0.87
        j54 = 0.5
        j55 = float(self.j55i2.get()) - float(self.j55i1.get())
        j57 = 0.12
        j59 =((float(self.j52i.get())* j53 * j54 * j55 / float(self.j58i.get()) ) + (float(self.j56i.get()) * j57 * j55 /float(self.j58i.get()) ) )/859.84

        self.o_59 = Label(tab8,text = "Power Required " +"                       " + str(j59),font="Verdana 10 bold",fg="#ffff00",bg="#404149")
        self.o_59.place(x=100, y= 500)

        '''c16i = my_sheet['E3']
        c16i.value = "Total Requirement"
        c16ii = my_sheet['H3']
        c16ii.value = str(g29)
        c17i = my_sheet['E4']
        c17i.value = " Cooling Water Requirement"
        c17ii = my_sheet['H4']
        c17ii.value = str(d11)
        c18i = my_sheet['E5']
        c18i.value = "Tank Capacity required"
        c18ii = my_sheet['H5']
        c18ii.value = str(o34)
        c19i = my_sheet['E6']
        c19i.value = "Power Required"
        c19ii = my_sheet['H6']
        c19ii.value = str(j59)

        my_wb.save("sample_data3.xlsx")'''
mywin = f1(tab8)
#tab9
##################################################################################
'''Label(tab9,text="",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=0,sticky=W)
Label(tab9,text="Site elevation",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=1,column=0,sticky=W)
Label(tab9,text="m",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=1,column=2,sticky=W)
Label(tab9,text="Friction Losses in Horizontal pipe, fittings and Piping components like NRV, strainer etc. ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2,sticky=W)
Label(tab9,text="MWC ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2,column=2,sticky=W)
Label(tab9,text="TG deck Elevation",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=0,sticky=W)
Label(tab9,text="M",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=2,sticky=W)
Label(tab9,text="Turbine Centerline",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4,column=0,sticky=W)
Label(tab9,text="M",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4,column=2,sticky=W)
Label(tab9,text="Lube Oil Skid floor Elevation",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=0,sticky=W)
Label(tab9,text="M",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=2,sticky=W)
Label(tab9,text="Net positive suction head Required as recommended by Shaft driven pump manufacturer. ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=0,sticky=W)
Label(tab9,text="MWC ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=2,sticky=W)
Label(tab9,text="Actual  elevation rasied  of Lube oil system after deducting equipment Pedestal ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=0,sticky=W)
Label(tab9,text="MT ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=2,sticky=W)
Label(tab9,text="",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=8,column=1,sticky=W)
c4 = Entry(tab9)
c13 = Entry(tab9)
c14 = Entry(tab9)
c15 =Entry(tab9)
c16 = Entry(tab9)
c22 = Entry(tab9)
c27 = Entry(tab9)
c4.grid(row=1,column=1)
c13.grid(row=2,column=1)
c14.grid(row=3,column=1)
c15.grid(row=4,column=1)
c16.grid(row=5,column=1)
c22.grid(row=6,column=1)
c27.grid(row=7,column=1)
c4.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c13.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c14.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c15.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c16.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c22.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c27.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
 
def function1():
    num1 = float(c4.get())
    k4 = 101325
    k5 = 288.16
    k6 = 9.80665
    k7 = 0.02896968
    k8 = 8.314462618
    h2 = k7*k6
    h3 = k5*k8
    i3 = -1*(h2/h3)
    f5 = i3*num1
    f6 = math.exp(f5)
    f7 = k4*f6
    f9 = f7*0.0075
    f10 = f9*13.663/1000
    c6 = f10
    c7 = -20
    c8 = -1*c7*13.663/1000
    c9 = c6 - c8
    num2 = float(c13.get())
    num3 = float(c14.get())
    num4 = float(c15.get())
    num5 = float(c16.get())
    c17 = num3+num4-num5
    c18 = c17 + num2
    c20 = c9-c18
    num6 = float(c22.get())
    c24 = num6+1
    num7 = float(c27.get())
    c32 = num6
    c34 = c20 + num7
    c35 = c20
    c37 = round(c35 - c32,2)
    c38 = c34 - c32
    Label(tab9,text="Condition : NPSH (Available ) 1 Meter higher than NPSH (Required) "+" "+str(c37),font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=12,column=0)
b = Button(tab9, text="Calculate", command=function1,bg="blue",fg="white")
b.grid(row=10, column=0, columnspan=2, rowspan=2)'''

#Pressure Drop Calc across Strainer
#################################################################################################


#New NPSH
################################################################
Label(tab11,text="",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=0,sticky=W)
Label(tab11,text="Site elevation",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=1,column=0,sticky=W)
Label(tab11,text="m",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=1,column=2,sticky=W)
Label(tab11,text="Friction Losses in Horizontal pipe, fittings and Piping components like NRV, strainer etc. ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2,sticky=W)
Label(tab11,text="MWC ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=2,column=2,sticky=W)
Label(tab11,text="TG deck Elevation",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=0,sticky=W)
Label(tab11,text="M",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=3,column=2,sticky=W)
Label(tab11,text="Turbine Centerline",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4,column=0,sticky=W)
Label(tab11,text="M",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=4,column=2,sticky=W)
Label(tab11,text="Lube Oil Skid floor Elevation",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=0,sticky=W)
Label(tab11,text="M",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=5,column=2,sticky=W)
Label(tab11,text="Net positive suction head Required as recommended by Shaft driven pump manufacturer. ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=0,sticky=W)
Label(tab11,text="MWC ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=6,column=2,sticky=W)
Label(tab11,text="Actual  elevation rasied  of Lube oil system after deducting equipment Pedestal ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=0,sticky=W)
Label(tab11,text="MT ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=7,column=2,sticky=W)
Label(tab11,text="",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=8,column=1,sticky=W)
c4 = Entry(tab11)
c13 = Entry(tab11)
c14 = Entry(tab11)
c15 =Entry(tab11)
c16 = Entry(tab11)
c22 = Entry(tab11)
c27 = Entry(tab11)
c4.grid(row=1,column=1)
c13.grid(row=2,column=1)
c14.grid(row=3,column=1)
c15.grid(row=4,column=1)
c16.grid(row=5,column=1)
c22.grid(row=6,column=1)
c27.grid(row=7,column=1)
c4.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c13.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c14.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c15.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c16.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c22.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
c27.config(bg="#F0E68C",fg="red",font="Verdana 10 bold")
 
def function1():
    num1 = float(c4.get())
    k4 = 101325
    k5 = 288.16
    k6 = 9.80665
    k7 = 0.02896968
    k8 = 8.314462618
    h2 = k7*k6
    h3 = k5*k8
    i3 = -1*(h2/h3)
    f5 = i3*num1
    f6 = math.exp(f5)
    f7 = k4*f6
    f9 = f7*0.0075
    f10 = f9*13.663/1000
    c6 = f10
    c7 = -20
    c8 = -1*c7*13.663/1000
    c9 = c6 - c8
    num2 = float(c13.get())
    num3 = float(c14.get())
    num4 = float(c15.get())
    num5 = float(c16.get())
    c17 = num3+num4-num5
    c18 = c17 + num2
    c20 = c9-c18
    num6 = float(c22.get())
    c24 = num6+1
    num7 = float(c27.get())
    c32 = num6
    c34 = c20 + num7
    c35 = c20
    c37 = round(c35 - c32,2)
    c38 = c34 - c32
    Label(tab11,text="NPSH Required by pump vendor"+" ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=12,column=0,sticky=W)
    a= Label(tab11,text=str(c32)+" "+"MWC",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
    #.grid(row=12,column=1)
    a.place(x=661,y=225)
    Label(tab11,text="NPSH Available with Pedestal",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=13,column=0,sticky=W)
    Label(tab11,text=str(c34)+" "+"MWC",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=13,column=1)
    Label(tab11,text="NPSH Available Without Pedestal",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=14,column=0,sticky = W)
    Label(tab11,text=str(c35)+" "+"MWC",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=14,column=1)
    Label(tab11,text="Condition : NPSH (Available ) 1 Meter higher than NPSH (Required) ",font="Verdana 10 bold",fg="#ffff00",bg="#404149").grid(row=15,column=0,sticky=W)
    b = Label(tab11,text=str(c37)+" "+"MWC",font="Verdana 10 bold",fg="#ffff00",bg="#404149")
    b.place(x=661,y=290)
b = Button(tab11, text="Calculate", command=function1,bg="blue",fg="white")
b.grid(row=10, column=0, columnspan=2, rowspan=2)

root.mainloop()